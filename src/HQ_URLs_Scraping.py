import serpapi
import os
from openpyxl import load_workbook
from dotenv import load_dotenv

#This script uses Serpapi. The API key is stored seperaperly in a .env file 
load_dotenv(dotenv_path="key.env")

api_key = os.getenv("SERPAPI_KEY")
client = serpapi.Client(api_key=api_key)


def get_URL_HQaddress(company_name):
#Start searching the company you want with Google
        result = client.search(
        q = company_name,
        engine = "google"
        )
#Get company's official website's url from the organic result
        organic_results = result.get("organic_results", [])
        if organic_results:
            url = organic_results[0]["link"]
        else:
            url = "Not Found"

#Get company's HQ address. This step requires multiple try statements, as the HQ address information may not always be found in the same location on the web
        try:
            knowledge_graph = result["knowledge_graph"]
            try:
                HQaddress = knowledge_graph["headquarters"]
            except KeyError:
                try:
                    HQaddress = knowledge_graph["founded_links"][0]["text"]
                except(KeyError, IndexError):
                    try:
                        HQaddress = knowledge_graph["address"]
                    except KeyError:
                        HQaddress = "Not Found"
        except KeyError:
            HQaddress = "Not Found"
        
        return url, HQaddress

#Input the url and HQaddress into your Excel file (xlsx file in this example)
book = load_workbook("Account Match.xlsx")
sheet = book.active

for row in range(2,sheet.max_row + 1):
    company_name = sheet[f"A{row}"].value
    
    if not sheet[f"A{row}"].value:
        break
            
    if sheet[f"B{row}"].value and sheet[f"C{row}"].value:
        print(f"Row {row} already processed. Skipping...")
        continue
        
    url, HQaddress = get_URL_HQaddress(company_name)
    sheet[f"B{row}"] = HQaddress
    sheet[f"C{row}"] = url
    
    book.save("Account Match.xlsx")

#Feel free to review the completed Account Match.xlsx file to see the results